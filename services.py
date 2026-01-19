# services.py
# ============================================================
# BUSINESS LOGIC FOR PHARM-POS SUPPLIER AGGREGATOR
# ============================================================
import io
import base64
import json
import re
from datetime import datetime, time, timedelta
from typing import List, Optional, Dict, Tuple, Set, Any
from collections import defaultdict
from typing import Dict, Tuple, Optional, List, Set, cast
import requests
import xmltodict
import chardet
from sqlalchemy.orm import Session
from sqlalchemy import text

import asyncio
import httpx
from sqlalchemy import func, case, cast, Numeric
from core import (
    HTTP_TIMEOUT,
    MAX_PRODUCTS_PER_PROVIDER,
)
import csv
from models import (
    Supplier,
    SupplierMapping,
    HourlyProduct,
    DailyProduct,
    CityResponse,
    ProductCompare,
    ProductCanonical,
    BarcodeAlias,
)

from repositories import (
    SupplierRepo,
    MappingRepo,
    HourlyRepo,
    DailyRepo,
    SupplierCityRepo,
    SupplierUnitRepo,
    CityResponseRepo,
    ProductCanonicalRepo,
    BarcodeAliasRepo,
    SupplierSrokRepo
)


from datetime import datetime
from decimal import Decimal
from typing import Optional, Dict, Tuple, cast as tcast
from uuid import UUID

from sqlalchemy.orm import Session

from stock_movement_model import StockMovement, StockMovementType
from repositories import StockMovementRepo
from repositories import HourlyRepo, BarcodeAliasRepo  # используем твои эталонные репо

from utils import normalize_numeric
from schemas import AggregatedItem, SupplierMappingCreate
from utils import nested_get, normalize_barcode, clean_unit, normalize_name, normalize_srok
from database import SessionLocal


from datetime import date


class SupplierSrokService:
    """
    Работает по логике:
    - в supplier_srok_response хранится маска для поставщика (например 'yyyymmdd')
    - сюда приходит реальный срок от API (например '20261031')
    - мы нормализуем его в 'dd-mm-yyyy' через normalize_srok(raw, pattern)
    """

    @staticmethod
    def resolve_srok(
        db: Session,
        provider_name: str,
        provider_srok_raw: str | None,
    ) -> str | None:

        if not provider_srok_raw:
            return None

        config = SupplierSrokRepo.get_by_provider(db, provider_name)
        if not config:
            # если формат не настроен — отдаём как есть
            return provider_srok_raw

        pattern = config.provider_srok_raw  # например 'yyyymmdd'
        return normalize_srok(provider_srok_raw, pattern)


def sanitize_xml(raw: str) -> str:
    end_tag = "</root>"
    idx = raw.lower().find(end_tag)
    if idx != -1:
        return raw[: idx + len(end_tag)]
    return raw


def clean_xml(raw: str) -> str:
    """
    Убирает мусор после </root>, который ломает xmltodict
    (часто встречается у rauza)
    """
    if not raw:
        return raw

    lower = raw.lower()
    end = lower.rfind("</root>")
    if end != -1:
        return raw[: end + len("</root>")]

    return raw

def sanitize_barcodes(raw) -> list[str]:
    """
    Оставляем ТОЛЬКО реальные баркоды:
    - только цифры
    - длина 6–14
    """
    cleaned = []

    for b in raw or []:
        try:
            b = str(b).strip()
        except Exception:
            continue

        if b.isdigit() and 6 <= len(b) <= 14:
            cleaned.append(b)

    return cleaned



# ============================================================
# CITY SERVICE
# ============================================================

class CityService:

    @staticmethod
    def extract_city_from_response(mapping: SupplierMapping, parsed_data: dict) -> Optional[str]:
        value = None

        # 1) city_path
        if mapping.city_path:
            value = nested_get(parsed_data, mapping.city_path)

        # 2) city / City
        if not value:
            value = parsed_data.get("city") or parsed_data.get("City")

        # 3) params.city
        if not value and isinstance(parsed_data, dict):
            params = parsed_data.get("params")
            if isinstance(params, dict):
                value = params.get("city")

        return str(value).strip() if value else None

    @staticmethod
    def resolve_city(db: Session, provider_name: str, city_code: str, city_name: str) -> str:
        city = (
            db.query(CityResponse)
            .filter(
                CityResponse.provider_name == provider_name,
                CityResponse.supplier_city_code == city_code
            )
            .first()
        )

        if city:
            return city.normalized_city

        try:
            obj = CityResponse(
                provider_name=provider_name,
                supplier_city_code=city_code,
                supplier_city_name=city_name,
                normalized_city=city_name.lower()
            )
            db.add(obj)
            db.commit()
            db.refresh(obj)
            return obj.normalized_city

        except Exception:
            db.rollback()

            # Дубликат? Значит создал кто-то другой
            city = (
                db.query(CityResponse)
                .filter(
                    CityResponse.provider_name == provider_name,
                    CityResponse.supplier_city_code == city_code
                )
                .first()
            )
            if city:
                return city.normalized_city

            raise

# ============================================================
# FETCH SERVICE (СИНХРОННЫЙ)
# ============================================================

class FetchService:
    """
    Transport-layer service.

    HTTP/HTTPS:
      - returns decoded TEXT (for json/xml parsing)

    FTP/SFTP:
      - returns RAW BYTES (for csv/excel parsing)
    """

    @staticmethod
    def fetch_http_text(
        url: str,
        login: str | None = None,
        password: str | None = None,
        *,
        verify_ssl: bool = False,
    ) -> str:
        """
        HTTP fetch that returns TEXT with robust encoding detection.
        Mirrors your previous logic but is explicit about HTTP.
        """
        if not url:
            raise ValueError("HTTP URL is empty")

        if not url.startswith(("http://", "https://")):
            url = "http://" + url

        headers = {"User-Agent": "PharmPOS/1.0"}

        if login and password:
            raw = f"{login}:{password}".encode("utf-8")
            headers["Authorization"] = "Basic " + base64.b64encode(raw).decode("ascii")

        response = requests.get(
            url,
            headers=headers,
            timeout=HTTP_TIMEOUT,
            verify=verify_ssl,
        )
        response.raise_for_status()

        raw_bytes = response.content
        encoding = chardet.detect(raw_bytes).get("encoding") or "utf-8"

        def looks_broken(txt: str) -> bool:
            bad = ["Ð", "Ñ", "â", "€", "", ""]
            return sum(1 for p in bad if p in txt) >= 2

        # autodetect
        try:
            txt = raw_bytes.decode(encoding, errors="strict")
            if not looks_broken(txt):
                return txt
        except Exception:
            pass

        # cp1251
        try:
            txt = raw_bytes.decode("cp1251", errors="strict")
            if not looks_broken(txt):
                return txt
        except Exception:
            pass

        # fallback
        return raw_bytes.decode("latin-1", errors="ignore")

    @staticmethod
    def fetch_ftp_file_bytes(supplier: Supplier) -> bytes:
        """
        Downloads a file from FTP/SFTP using Supplier fields:
          ftp_host, ftp_port, ftp_login, ftp_password, ftp_path, ftp_type

        Returns raw bytes (for csv/excel parsing).
        """
        ftp_host = (supplier.ftp_host or "").strip()
        ftp_path = (supplier.ftp_path or "").strip()
        ftp_type = (supplier.ftp_type or "").strip().lower()
        ftp_port = int(supplier.ftp_port or 21)

        if not ftp_host:
            raise ValueError("Supplier.ftp_host is empty")
        if not ftp_path:
            raise ValueError("Supplier.ftp_path is empty")
        if ftp_type not in {"ftp", "sftp"}:
            raise ValueError("Supplier.ftp_type must be 'ftp' or 'sftp'")

        if ftp_type == "ftp":
            return FetchService._download_via_ftp(
                host=ftp_host,
                port=ftp_port,
                username=supplier.ftp_login,
                password=supplier.ftp_password,
                path=ftp_path,
            )

        # sftp
        return FetchService._download_via_sftp(
            host=ftp_host,
            port=ftp_port if ftp_port else 22,
            username=supplier.ftp_login,
            password=supplier.ftp_password,
            path=ftp_path,
        )

    @staticmethod
    def _download_via_ftp(
        *,
        host: str,
        port: int,
        username: Optional[str],
        password: Optional[str],
        path: str,
    ) -> bytes:
        import ftplib

        bio = io.BytesIO()

        # NOTE: timeout uses HTTP_TIMEOUT for consistency
        ftp = ftplib.FTP()
        ftp.connect(host=host, port=port, timeout=HTTP_TIMEOUT)

        user = username or "anonymous"
        pwd = password or ""
        ftp.login(user=user, passwd=pwd)

        # Passive mode is typical for NAT environments
        try:
            ftp.set_pasv(True)
        except Exception:
            pass

        # Retrieve the file
        try:
            ftp.retrbinary(f"RETR {path}", bio.write)
        finally:
            try:
                ftp.quit()
            except Exception:
                try:
                    ftp.close()
                except Exception:
                    pass

        return bio.getvalue()

    @staticmethod
    def _download_via_sftp(
        *,
        host: str,
        port: int,
        username: Optional[str],
        password: Optional[str],
        path: str,
    ) -> bytes:
        """
        Uses paramiko if available. If not installed, raises a clear error.
        """
        try:
            import paramiko  # type: ignore
        except Exception as e:
            raise RuntimeError(
                "SFTP requested but 'paramiko' is not installed. "
                "Install paramiko to enable SFTP support."
            ) from e

        if not username or not password:
            raise ValueError("SFTP requires ftp_login and ftp_password on Supplier")

        transport = None
        sftp = None

        try:
            transport = paramiko.Transport((host, port or 22))
            transport.connect(username=username, password=password)

            sftp = paramiko.SFTPClient.from_transport(transport)
            if not sftp:
                raise ValueError("Not found any transport")

            with sftp.open(path, "rb") as f:
                return f.read()

        finally:
            try:
                if sftp is not None:
                    sftp.close()
            except Exception:
                pass
            try:
                if transport is not None:
                    transport.close()
            except Exception:
                pass
# ============================================================
# FETCH SERVICE (АСИНХРОННЫЙ — ТОЛЬКО HTTP)
# ============================================================



# ============================================================
# PARSE SERVICE — FINAL VERSION
# ============================================================

class ParseService:
    """
    ParseService отвечает ТОЛЬКО за:
    - парсинг входных данных по формату
    - получение списка items
    - нормализацию одного item по mapping

    НЕ:
    - не ходит в БД
    - не знает про HTTP / FTP
    - не знает про hourly / daily
    """

    # ---------------------------------------------------------
    # MAIN PARSE ENTRYPOINT
    # ---------------------------------------------------------
    @staticmethod
    def parse(
        raw: str | bytes,
        data_format: str,
    ) -> dict | List[dict]:
        """
        Возвращает:
        - dict            → json / xml
        - list[dict]      → csv / excel
        """
        fmt = (data_format or "").lower().strip()

        if fmt == "json":
            if not isinstance(raw, str):
                raise ValueError("JSON expects text input")
            return json.loads(raw)

        if fmt == "xml":
            if not isinstance(raw, str):
                raise ValueError("XML expects text input")
            return xmltodict.parse(raw)

        if fmt == "csv":
            if not isinstance(raw, (bytes, bytearray)):
                raise ValueError("CSV expects bytes input")
            return ParseService._parse_csv(raw)

        if fmt == "excel":
            if not isinstance(raw, (bytes, bytearray)):
                raise ValueError("Excel expects bytes input")
            return ParseService._parse_excel(raw)

        raise ValueError(f"Unsupported format: {data_format}")

    # ---------------------------------------------------------
    # CSV
    # ---------------------------------------------------------
    @staticmethod
    def _parse_csv(raw: bytes) -> List[dict]:
        """
        CSV → list[dict]
        """
        text = raw.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))

        rows: List[dict] = []
        for row in reader:
            if not row:
                continue
            rows.append({k.strip(): v for k, v in row.items()})

        return rows

    # ---------------------------------------------------------
    # EXCEL (xlsx)
    # ---------------------------------------------------------
    @staticmethod
    def _parse_excel(raw: bytes) -> List[dict]:
        """
        Excel (.xlsx) → list[dict]
        """
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError(
                "Excel parsing requires openpyxl"
            ) from e

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        if not ws:
            raise
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() for h in rows[0]]

        items: List[dict] = []
        for r in rows[1:]:
            row = {}
            for i, value in enumerate(r):
                if i < len(headers):
                    row[headers[i]] = value
            items.append(row)

        return items

    # ---------------------------------------------------------
    # ITEMS EXTRACTION
    # ---------------------------------------------------------
    @staticmethod
    def get_items_by_path(
        parsed_data: dict | List[dict],
        items_path: Optional[str],
    ) -> List[dict]:
        """
        Универсальный extractor:
        - json / xml → items_path
        - csv / excel → parsed_data уже list[dict]
        """
        if isinstance(parsed_data, list):
            return parsed_data

        if not items_path:
            return []

        current: Any = parsed_data

        for part in items_path.split("."):
            if isinstance(current, list):
                if not current:
                    return []
                current = current[0]

            if not isinstance(current, dict):
                return []

            current = current.get(part)
            if current is None:
                return []

        if isinstance(current, list):
            return current

        if isinstance(current, dict):
            return [current]

        return []

    # ---------------------------------------------------------
    # NORMALIZE ONE ITEM
    # ---------------------------------------------------------
    @staticmethod
    def normalize_item(
        item: dict,
        mapping: SupplierMapping,
    ) -> dict:
        """
        Приводит сырой item к унифицированной структуре
        (БЕЗ бизнес-логики: unit, srok, city и т.д.)
        """

        def get_value(key: Optional[str]):
            if not key:
                return None
            if not isinstance(item, dict):
                return None
            return item.get(key)

        return {
            "sku_uid": get_value(mapping.sku_uid),
            "sku_name": get_value(mapping.sku_name),
            "sku_price": get_value(mapping.sku_price),
            "sku_stock": get_value(mapping.sku_stock),

            "sku": get_value(mapping.sku),
            "sku_serial": get_value(mapping.sku_serial),

            "sku_barcodes": get_value(mapping.sku_barcodes),
            "sku_srok": get_value(mapping.sku_srok),
            "sku_step": get_value(mapping.sku_step),
            "sku_marker": get_value(mapping.sku_marker),
            "sku_pack": get_value(mapping.sku_pack),
            "sku_box": get_value(mapping.sku_box),

            "unit": get_value(mapping.unit),
            "min_order": get_value(mapping.min_order),

            "producer": get_value(mapping.producer),
            "producer_country": get_value(mapping.producer_country),
        }

# ============================================================
# SYNC SERVICE
# ============================================================

class SyncService:

    DEFAULT_UNIT = "778"  # упаковка

    _unit_cache: Dict[Tuple[str, str], str] = {}

    # ---------------------------------------------------------
    # UNIT CACHE
    # ---------------------------------------------------------
    @staticmethod
    def _reset_unit_cache():
        SyncService._unit_cache.clear()

    @staticmethod
    def _normalize_unit_cached(
        db: Session,
        provider: str,
        raw_unit: Optional[str],
    ) -> str:
        key = (provider, str(raw_unit) if raw_unit else "")
        cached = SyncService._unit_cache.get(key)
        if cached is not None:
            return cached

        if not raw_unit:
            value = SyncService.DEFAULT_UNIT
        else:
            fixed = SupplierUnitRepo.find(db, provider, raw_unit)
            value = fixed.normalized_unit if fixed else SyncService.DEFAULT_UNIT

        SyncService._unit_cache[key] = value
        return value

    # ---------------------------------------------------------
    # TRANSPORT SELECTOR
    # ---------------------------------------------------------
    @staticmethod
    def _fetch_supplier_data(
        supplier: Supplier,
        mapping: SupplierMapping,
        *,
        city_code: Optional[str] = None,
    ) -> str | bytes:
        """
        Выбирает источник:
        - FTP/SFTP → bytes
        - HTTP/HTTPS → text
        """
        if supplier.ftp_host and supplier.ftp_path:
            return FetchService.fetch_ftp_file_bytes(supplier)

        # HTTP
        base_url = (
            supplier.json_url_get_price
            if mapping.format == "json"
            else supplier.xml_url_get_price
        )

        if not base_url:
            raise RuntimeError("Price URL not configured")

        if city_code:
            if "{city}" in base_url:
                url = base_url.format(city=city_code)
            else:
                param = supplier.city_param_name or "city_id"
                sep = "&" if "?" in base_url else "?"
                url = f"{base_url}{sep}{param}={city_code}"
        else:
            url = base_url

        return FetchService.fetch_http_text(url, supplier.login, supplier.password)

    # ---------------------------------------------------------
    # SINGLE SUPPLIER SYNC
    # ---------------------------------------------------------
    @staticmethod
    def sync_single_supplier(db: Session, supplier: Supplier) -> dict:

        mapping = MappingRepo.get_by_provider(db, supplier.provider_name)
        if not mapping:
            return {
                "provider": supplier.provider_name,
                "status": "failed",
                "message": "Mapping not found",
            }

        total_products = 0
        results = []

        # -----------------------------------------------------
        # MULTI CITY
        # -----------------------------------------------------
        if mapping.city_in_params:
            cities = (
                db.query(CityResponse)
                .filter(CityResponse.provider_name == supplier.provider_name)
                .all()
            )

            if not cities:
                return {
                    "provider": supplier.provider_name,
                    "status": "error",
                    "message": "No cities registered",
                }

            for city in cities:
                try:
                    raw = SyncService._fetch_supplier_data(
                        supplier,
                        mapping,
                        city_code=city.supplier_city_code,
                    )

                    parsed = ParseService.parse(raw, mapping.format)
                    items = ParseService.get_items_by_path(
                        parsed,
                        mapping.items_path,
                    )

                    objects: List[HourlyProduct] = []

                    for item in items[:MAX_PRODUCTS_PER_PROVIDER]:
                        normalized = ParseService.normalize_item(item, mapping)

                        barcodes = [
                            b for b in normalize_barcode(normalized.get("sku_barcodes"))
                            if b.isdigit() and 6 <= len(b) <= 14
                        ]
                        if not barcodes:
                            continue

                        normalized["sku_barcodes"] = barcodes
                        normalized["unit"] = SyncService._normalize_unit_cached(
                            db, supplier.provider_name, normalized.get("unit")
                        )
                        normalized["sku_srok"] = SupplierSrokService.resolve_srok(
                            db=db,
                            provider_name=supplier.provider_name,
                            provider_srok_raw=normalized.get("sku_srok"),
                        )

                        objects.append(
                            HourlyProduct(
                                provider_id=supplier.id,
                                provider_name=supplier.provider_name,
                                provider_bin=supplier.provider_bin,
                                city=city.normalized_city,
                                **normalized,
                            )
                        )

                    if objects:
                        HourlyRepo.bulk_create(db, objects)

                    total_products += len(objects)

                    results.append({
                        "city": city.normalized_city,
                        "processed": len(objects),
                        "status": "success",
                    })

                except Exception as e:
                    db.rollback()
                    results.append({
                        "city": city.normalized_city,
                        "status": "error",
                        "message": str(e),
                    })

        # -----------------------------------------------------
        # SINGLE CITY
        # -----------------------------------------------------
        else:
            try:
                raw = SyncService._fetch_supplier_data(supplier, mapping)
                parsed = ParseService.parse(raw, mapping.format)

                items = ParseService.get_items_by_path(
                    parsed,
                    mapping.items_path,
                )

                objects: List[HourlyProduct] = []

                for item in items[:MAX_PRODUCTS_PER_PROVIDER]:
                    normalized = ParseService.normalize_item(item, mapping)

                    barcodes = [
                        b for b in normalize_barcode(normalized.get("sku_barcodes"))
                        if b.isdigit()
                    ]
                    if not barcodes:
                        continue

                    normalized["sku_barcodes"] = barcodes
                    normalized["unit"] = SyncService._normalize_unit_cached(
                        db, supplier.provider_name, normalized.get("unit")
                    )
                    normalized["sku_srok"] = SupplierSrokService.resolve_srok(
                        db=db,
                        provider_name=supplier.provider_name,
                        provider_srok_raw=normalized.get("sku_srok"),
                    )

                    objects.append(
                        HourlyProduct(
                            provider_id=supplier.id,
                            provider_name=supplier.provider_name,
                            provider_bin=supplier.provider_bin,
                            city=None,
                            **normalized,
                        )
                    )

                if objects:
                    HourlyRepo.bulk_create(db, objects)

                total_products = len(objects)

                results.append({
                    "city": None,
                    "processed": total_products,
                    "status": "success",
                })

            except Exception as e:
                db.rollback()
                return {
                    "provider": supplier.provider_name,
                    "status": "error",
                    "message": str(e),
                }

        return {
            "provider": supplier.provider_name,
            "status": "success",
            "total_products": total_products,
            "cities_processed": len(results),
            "details": results,
        }

    # ---------------------------------------------------------
    # DAILY SNAPSHOT
    # ---------------------------------------------------------
    @staticmethod
    def run_daily_snapshot(db: Session) -> int:
        hourly = HourlyRepo.get_all_after_23(db)
        daily: List[DailyProduct] = []

        for h in hourly:
            daily.append(
                DailyProduct(
                    provider_id=h.provider_id,
                    provider_name=h.provider_name,
                    provider_bin=h.provider_bin,
                    city=h.city,
                    producer=h.producer,
                    producer_country=h.producer_country,
                    sku_uid=h.sku_uid,
                    sku_name=h.sku_name,
                    sku_price=h.sku_price,
                    sku_stock=h.sku_stock,
                    sku=h.sku,
                    sku_serial=h.sku_serial,
                    sku_barcodes=h.sku_barcodes,
                    sku_srok=h.sku_srok,
                    sku_step=h.sku_step,
                    sku_marker=h.sku_marker,
                    sku_pack=h.sku_pack,
                    sku_box=h.sku_box,
                    unit=h.unit,
                    min_order=h.min_order,
                    snapshot_date=datetime.utcnow().date(),
                )
            )

        if daily:
            DailyRepo.bulk_create(db, daily)

        return len(daily)

    # ---------------------------------------------------------
    # HOURLY ENTRYPOINT
    # ---------------------------------------------------------
    @staticmethod
    def run_hourly_sync(db: Session):
        suppliers = SupplierRepo.get_active(db)
        if not suppliers:
            return {"status": "error", "message": "No active suppliers"}

        SyncService._reset_unit_cache()

        results = []
        total = 0

        for supplier in suppliers:
            res = SyncService.sync_single_supplier(db, supplier)
            results.append(res)
            total += res.get("total_products", 0)

        return {
            "status": "success",
            "total_suppliers": len(suppliers),
            "total_products": total,
            "details": results,
        }

    # ---------------------------------------------------------
    # CLEANUP
    # ---------------------------------------------------------
    @staticmethod
    def cleanup_hourly_table(db: Session) -> bool:
        HourlyRepo.clear_table(db)
        return True

def to_str(val):
    return str(val) if val is not None else None

# ============================================================
# PRODUCT SERVICE
# ============================================================

class ProductService:

    @staticmethod
    def get_by_barcode(db: Session, barcode: str, city: Optional[str] = None, provider_name: Optional[str] = None) -> List[AggregatedItem]:

        # 1) HOURLY
        items = HourlyRepo.get_latest_by_barcode(db, barcode, city=city, provider_name=provider_name)

        # 2) DAILY fallback
        if not items:
            items = DailyRepo.get_latest_by_barcode(db, barcode, city=city, provider_name=provider_name)

        result: List[AggregatedItem] = []

        for item in items:

            matched = None
            if item.sku_barcodes:
                for b in item.sku_barcodes:
                    if str(b).strip() == str(barcode).strip():
                        matched = b
                        break

            if not matched:
                continue

            result.append(
                AggregatedItem(
                    provider_name=item.provider_name,
                    provider_bin=item.provider_bin,
                    city=item.city,
                    producer=item.producer,
                    producer_country=item.producer_country,
                    sku_uid=item.sku_uid,
                    sku_name=item.sku_name,
                    sku_barcode=matched,
                    sku_price=to_str(item.sku_price),
                    sku_stock=to_str(item.sku_stock),
                    sku_step=item.sku_step,
                    unit=item.unit,
                    min_order=item.min_order,
                    sku_serial=item.sku_serial,
                    sku_srok=item.sku_srok,
                    sku_marker=item.sku_marker,
                    last_update=item.created_at
                )
            )

        return result


# ============================================================
# MAPPING CREATE + AUTO SYNC
# ============================================================

class SupplierMappingService:

    @staticmethod
    def create_mapping(db: Session, mapping_data: SupplierMappingCreate) -> dict:

        mapping = SupplierMapping(**mapping_data.dict())
        db.add(mapping)
        db.commit()
        db.refresh(mapping)

        supplier = SupplierRepo.get_by_name(db, mapping.provider_name)

        if not supplier:
            return {
                "status": "warning",
                "message": "Mapping created, but supplier not found",
                "mapping_id": str(mapping.id)
            }

        result = SyncService.sync_single_supplier(db, supplier)

        return {
            "status": "success",
            "mapping_id": str(mapping.id),
            "sync_result": result,
        }


# ============================================================
# UNIVERSAL ANALYTICS SERVICE — DAILY / WEEKLY / MONTHLY
# ============================================================

class AnalyticsService:

    # ------------------------------------------------------------
    @staticmethod
    def get_period_bounds(period: str, start_date=None, end_date=None):
        """
        Возвращает (start_dt, end_dt)
        """
        now = datetime.utcnow()

        if period == "today":
            start = datetime.combine(now.date(), time.min)
            return start, now

        if period == "week":
            return now - timedelta(days=7), now

        if period == "month":
            return now - timedelta(days=30), now

        if period == "custom":
            if not start_date or not end_date:
                raise ValueError("Custom period requires start_date and end_date")
            return start_date, end_date

        raise ValueError(f"Unknown period: {period}")

    # ------------------------------------------------------------
    @staticmethod
    def get_hot_products_period(
        db: Session,
        period: str = "today",
        start_date=None,
        end_date=None,
        provider_name: Optional[str] = None,
        city: Optional[str] = None,
        limit: int = 10
    ) -> dict:
        """
        🔥 HOT PRODUCTS (корректная логика)

        Использует ТОЛЬКО stock_movements:
        - sold_qty      = сумма |delta| где delta < 0
        - restocked_qty = сумма delta где delta > 0
        - net_delta     = сумма delta
        """

        start_dt, end_dt = AnalyticsService.get_period_bounds(
            period, start_date, end_date
        )

        rows = StockMovementRepo.aggregate_for_period(
            db=db,
            start_dt=start_dt,
            end_dt=end_dt,
            provider_name=provider_name,
            city=city,
            limit=limit,
        )

        if not rows:
            return {
                "status": "empty",
                "period": period,
                "from": start_dt.isoformat(),
                "to": end_dt.isoformat(),
                "items": [],
            }

        items: List[Dict] = []

        for r in rows:
            sold = r["sold_qty"]
            restocked = r["restocked_qty"]
            net = r["net_delta"]

            # процент считаем ТОЛЬКО если были продажи
            percent = None
            if sold > 0:
                base = sold + max(restocked, 0)
                if base > 0:
                    percent = round((sold / base) * 100, 2)

            items.append({
                "provider_name": r["provider_name"],
                "city": r["city"],

                "canonical_id": r["canonical_id"],
                "sku_uid": r["sku_uid"],
                "sku_name": r["sku_name"],

                "sold": round(sold, 3),
                "restocked": round(restocked, 3),
                "net_delta": round(net, 3),

                "percent": percent,
            })

        # сортируем по реальному спросу
        items.sort(
            key=lambda x: (x["sold"], abs(x["net_delta"])),
            reverse=True,
        )

        return {
            "status": "success",
            "period": period,
            "from": start_dt.isoformat(),
            "to": end_dt.isoformat(),
            "items": items[:limit],
        }


# ============================================================
# CANONICAL RESOLVE SERVICE
# ============================================================
from uuid import UUID 
from sqlalchemy import select

def _row_get(row, key):
    try:
        return getattr(row, key)
    except Exception:
        try:
            return row._mapping.get(key)
        except Exception:
            return None
from repositories import PostProcessStateRepo
class CanonicalResolveService:
    """
    Строит и обновляет product_canonical и barcode_aliases
    на основе ПОСЛЕДНИХ записей hourly_products.
    """

    @staticmethod
    def rebuild_from_hourly(db: Session) -> None | datetime:
        state = PostProcessStateRepo.get(
            db
        )

        items = HourlyRepo.get_newer_than(
            db,
            state.last_hourly_at,
        )

        if not items:
            return  # ⬅️ skip, ничего нового

        # ---- дальше ТВОЯ СУЩЕСТВУЮЩАЯ ЛОГИКА ----
        # НИЧЕГО в алгоритме не меняем

        
        # после обработки
        last_at = items[-1].created_at

       
        # =========================
        # 2. preload maps
        # =========================
        canonical_by_name = ProductCanonicalRepo.preload_all(db)
        alias_map = BarcodeAliasRepo.preload_all(db)

        canonical_by_barcode: Dict[str, UUID] = {}
        rows = db.execute(
            select(
                ProductCanonical.canonical_barcode,
                ProductCanonical.id,
            )
        ).all()

        for r in rows:
            if r.canonical_barcode:
                canonical_by_barcode[str(r.canonical_barcode)] = r.id

        seen_keys: Set[Tuple[Optional[str], Optional[str], Tuple[str, ...]]] = set()
        new_aliases: List[Tuple[str, Optional[str], UUID]] = []

        # =========================
        # 3. PROCESS ITEMS
        # =========================
        for item in items:
            raw_barcodes = _row_get(item, "sku_barcodes") or []
            barcodes: List[str] = []

            for b in raw_barcodes:
                s = str(b).strip()
                if s.isdigit() and 6 <= len(s) <= 14:
                    barcodes.append(s)

            name_key = normalize_name(item.sku_name) if item.sku_name else None
            producer = item.producer

            if not barcodes and not name_key:
                continue

            combo_key = (name_key, producer, tuple(sorted(barcodes)))
            if combo_key in seen_keys:
                continue
            seen_keys.add(combo_key)

            canonical_id: Optional[UUID] = None

            # 1) alias
            for b in barcodes:
                if b in alias_map:
                    canonical_id = alias_map[b]
                    break

            # 2) canonical_barcode
            if not canonical_id:
                for b in barcodes:
                    if b in canonical_by_barcode:
                        canonical_id = canonical_by_barcode[b]
                        break

            # 3) name_key + producer
            if not canonical_id and name_key:
                canonical_id = canonical_by_name.get((name_key, producer))

            # =========================
            # CREATE CANONICAL
            # =========================
            if not canonical_id:
                main_barcode = barcodes[0] if barcodes else None

                obj = ProductCanonicalRepo.create(
                    db=db,
                    canonical_barcode=main_barcode,
                    name_key=name_key,
                    producer=producer,
                    producer_country=item.producer_country,
                )

                canonical_id = obj.id
                canonical_by_name[(name_key, producer)] = canonical_id
                if main_barcode:
                    canonical_by_barcode[main_barcode] = canonical_id

            # =========================
            # CREATE ALIASES
            # =========================
            for b in barcodes:
                if b not in alias_map:
                    new_aliases.append(
                        (b, item.provider_name, canonical_id)
                    )
                    alias_map[b] = canonical_id

        # =========================
        # 4. SAVE ALIASES
        # =========================
        for barcode, provider_name, canonical_id in new_aliases:
            BarcodeAliasRepo.get_or_create(
                db=db,
                provider_name=provider_name,
                barcode=barcode,
                canonical_id=canonical_id,
            )

        
        HourlyRepo.attach_canonical_ids(db)
     


        # коммит здесь безопасен и финальный
        
        return last_at

    @staticmethod
    def rebuild_full(db: Session) -> None:
        """
        NIGHTLY FULL CANONICAL REBUILD
        Safety-net: сбрасывает cursor и прогоняет ВСЕ hourly.
        """

        # 1️⃣ сбрасываем cursor
        PostProcessStateRepo.update_last_hourly_at(
            db,
            last_hourly_at=None,
        )

        # 2️⃣ запускаем обычный rebuild
        CanonicalResolveService.rebuild_from_hourly(db)

# ============================================================
# PRODUCT COMPARE SERVICE
# ============================================================


class ProductCompareService:

    @staticmethod
    def rebuild_from_hourly(db: Session) -> None:
        """
        Инкрементальный ORM-rebuild product_compare
        Работает ТОЛЬКО по изменённым canonical_id
        """

        state = PostProcessStateRepo.get(db)
        last_dt = state.last_hourly_at

        canonical_ids = HourlyRepo.get_changed_canonical_ids(
            db,
            last_hourly_at=last_dt,
        )
        if not canonical_ids:
            return

        rows = (
            db.query(
                ProductCanonical.id.label("canonical_id"),
                func.min(BarcodeAlias.barcode).label("barcode"),
                func.min(HourlyProduct.sku_name).label("sku_name"),

                func.max(
                    case(
                        (HourlyProduct.provider_name == "atamiras",
                         cast(func.replace(HourlyProduct.sku_price, ",", "."), Numeric)),
                        else_=None,
                    )
                ).label("price_atamiras"),

                func.max(
                    case(
                        (HourlyProduct.provider_name == "medservice",
                         cast(func.replace(HourlyProduct.sku_price, ",", "."), Numeric)),
                        else_=None,
                    )
                ).label("price_medservice"),

                func.max(
                    case(
                        (HourlyProduct.provider_name == "stopharm",
                         cast(func.replace(HourlyProduct.sku_price, ",", "."), Numeric)),
                        else_=None,
                    )
                ).label("price_stopharm"),

                func.max(
                    case(
                        (HourlyProduct.provider_name == "amanat",
                         cast(func.replace(HourlyProduct.sku_price, ",", "."), Numeric)),
                        else_=None,
                    )
                ).label("price_amanat"),

                func.max(
                    case(
                        (HourlyProduct.provider_name == "rauza",
                         cast(func.replace(HourlyProduct.sku_price, ",", "."), Numeric)),
                        else_=None,
                    )
                ).label("price_rauza"),
            )
            .join(BarcodeAlias, BarcodeAlias.canonical_id == ProductCanonical.id)
            .join(HourlyProduct, HourlyProduct.canonical_id == ProductCanonical.id)
            .filter(ProductCanonical.id.in_(canonical_ids))
            .group_by(ProductCanonical.id)
            .all()
        )

        existing = {
            pc.canonical_id: pc
            for pc in db.query(ProductCompare)
            .filter(ProductCompare.canonical_id.in_(canonical_ids))
            .all()
        }

        for r in rows:
            payload = dict(
                canonical_id=r.canonical_id,
                barcode=r.barcode,
                sku_name=r.sku_name,
                price_atamiras=str(r.price_atamiras) if r.price_atamiras else None,
                price_medservice=str(r.price_medservice) if r.price_medservice else None,
                price_stopharm=str(r.price_stopharm) if r.price_stopharm else None,
                price_amanat=str(r.price_amanat) if r.price_amanat else None,
                price_rauza=str(r.price_rauza) if r.price_rauza else None,
            )

            obj = existing.get(r.canonical_id)
            if obj:
                for k, v in payload.items():
                    setattr(obj, k, v)
            else:
                db.add(ProductCompare(**payload))
    @staticmethod
    def rebuild_full(db: Session) -> None:
        """
        NIGHTLY FULL REBUILD product_compare
        """

        db.execute(text("TRUNCATE TABLE product_compare"))
        db.commit()

        db.execute(
            text("""
                INSERT INTO product_compare (
                    canonical_id,
                    barcode,
                    sku_name,
                    price_atamiras,
                    price_medservice,
                    price_stopharm,
                    price_amanat,
                    price_rauza
                )
                SELECT
                    pc.id AS canonical_id,
                    MIN(b.code) AS barcode,
                    MIN(h.sku_name) AS sku_name,

                    MAX(CASE WHEN h.provider_name='atamiras'
                        AND h.sku_price ~ '^[0-9]+([.,][0-9]+)?$'
                        THEN REPLACE(h.sku_price, ',', '.')::numeric END)::text,

                    MAX(CASE WHEN h.provider_name='medservice'
                        AND h.sku_price ~ '^[0-9]+([.,][0-9]+)?$'
                        THEN REPLACE(h.sku_price, ',', '.')::numeric END)::text,

                    MAX(CASE WHEN h.provider_name='stopharm'
                        AND h.sku_price ~ '^[0-9]+([.,][0-9]+)?$'
                        THEN REPLACE(h.sku_price, ',', '.')::numeric END)::text,

                    MAX(CASE WHEN h.provider_name='amanat'
                        AND h.sku_price ~ '^[0-9]+([.,][0-9]+)?$'
                        THEN REPLACE(h.sku_price, ',', '.')::numeric END)::text,

                    MAX(CASE WHEN h.provider_name='rauza'
                        AND h.sku_price ~ '^[0-9]+([.,][0-9]+)?$'
                        THEN REPLACE(h.sku_price, ',', '.')::numeric END)::text

                FROM hourly_products h
                JOIN LATERAL jsonb_array_elements_text(h.sku_barcodes) b(code) ON TRUE
                JOIN barcode_aliases ba ON ba.barcode = b.code
                JOIN product_canonical pc ON pc.id = ba.canonical_id
                GROUP BY pc.id
            """)
        )

        db.commit()

class PostProcessService:
    """
    Оркестратор postprocess-логики.

    Гарантии:
    - инкрементальная обработка
    - согласованность canonical / stock_movements / product_compare
    - единая транзакция
    - корректный cursor
    """

    @staticmethod
    def rebuild_all(db: Session) -> Dict[str, int | str]:

        # --------------------------------------------
        # TRY LOCK
        # --------------------------------------------
        if not PostProcessStateRepo.try_set_running(db):
            return {
                "status": "skipped",
                "reason": "already running",
            }

        try:
            # ====================================================
            # 1️⃣ CANONICAL RESOLVE (incremental)
            # ====================================================
            last_hourly_at = CanonicalResolveService.rebuild_from_hourly(db)

            # если canonical ничего не делал — нет смысла идти дальше
            if not last_hourly_at:
                PostProcessStateRepo.set_success(
                    db,
                    last_hourly_at=PostProcessStateRepo.get(db).last_hourly_at,
                )
                db.flush()
                db.commit()
                return {
                    "status": "skipped",
                    "reason": "no new hourly data",
                }

            # ====================================================
            # 2️⃣ STOCK MOVEMENTS (incremental)
            # ====================================================
            total_movements = 0

            """suppliers = SupplierRepo.get_active(db)
            for supplier in suppliers:
                total_movements += StockMovementService.build_hourly_movements(
                    db,
                    provider_name=supplier.provider_name,
                )"""

            # ====================================================
            # 3️⃣ PRODUCT COMPARE (incremental)
            # ====================================================
            ProductCompareService.rebuild_from_hourly(db)

            # ====================================================
            # FINALIZE (ONE COMMIT)
            # ====================================================
            PostProcessStateRepo.set_success(
                db,
                last_hourly_at=last_hourly_at,
            )
            db.commit()

            return {
                "canonical": 1,
                "stock_movements": total_movements,
                "product_compare": 1,
            }

        except Exception:
            db.rollback()
            PostProcessStateRepo.set_failed(db)
            raise


class StockMovementService:
    """
    Генерирует stock_movements из hourly snapshots.
    Инкрементально, chunk-based, без full scan.
    """

    @staticmethod
    def _to_float(val) -> Optional[float]:
        if val is None:
            return None
        try:
            out = normalize_numeric(val)
            return float(out) if out is not None else None
        except Exception:
            return None

    KEYS_BATCH_SIZE = 1000  # ⬅️ ЭТАЛОН

    @staticmethod
    def _build_for_provider(
        db: Session,
        *,
        provider_name: str,
        chunk_size: int,
    ) -> int:
        cursor = StockMovementRepo.get_or_create_cursor(db, provider_name)
        since = cursor.last_hourly_processed_at

        total_created: int = 0
        last_processed_at: Optional[datetime] = since

        # ======================================================
        # ITERATE HOURLY CHUNKS
        # ======================================================
        for chunk in HourlyRepo.iter_newer_than_for_provider(
            db,
            provider_name=provider_name,
            since=since,
            chunk_size=chunk_size,
        ):
            if not chunk:
                continue

            # -----------------------------------------------
            # collect unique (canonical_id, city)
            # -----------------------------------------------
            keys: Set[Tuple[UUID, Optional[str]]] = set()

            for h in chunk:
                if h.canonical_id is None:
                    continue

                canonical_id = tcast(UUID, h.canonical_id)
                keys.add((canonical_id, h.city))

            last_processed_at = chunk[-1].created_at

            if not keys:
                continue

            # -----------------------------------------------
            # load previous snapshots (SUB-BATCHED)
            # -----------------------------------------------
            prev_cache: Dict[Tuple[UUID, Optional[str]], Optional[float]] = {}

            keys_list = list(keys)
            for i in range(0, len(keys_list), StockMovementService.KEYS_BATCH_SIZE):
                keys_batch = keys_list[i : i + StockMovementService.KEYS_BATCH_SIZE]

                prev_raw = HourlyRepo.get_prev_snapshots_batch(
                    db,
                    provider_name=provider_name,
                    keys=keys_batch,
                    before_dt=chunk[0].created_at,
                )

                for k, v in prev_raw.items():
                    prev_cache[k] = StockMovementService._to_float(v)

            # -----------------------------------------------
            # process chunk → stock_movements
            # -----------------------------------------------
            rows: List[dict] = []

            for h in chunk:
                if h.canonical_id is None:
                    continue

                canonical_id = tcast(UUID, h.canonical_id)
                key = (canonical_id, h.city)

                stock_after = StockMovementService._to_float(h.sku_stock)
                if stock_after is None:
                    continue

                stock_before = prev_cache.get(key)

                # first snapshot for this (canonical_id, city)
                if stock_before is None:
                    prev_cache[key] = stock_after
                    continue

                delta = stock_after - stock_before
                if delta == 0:
                    prev_cache[key] = stock_after
                    continue

                rows.append({
                    "provider_name": h.provider_name,
                    "city": h.city,
                    "canonical_id": canonical_id,
                    "sku_uid": h.sku_uid,
                    "sku_name": h.sku_name,
                    "stock_before": stock_before,
                    "stock_after": stock_after,
                    "delta": delta,
                    "movement_type": (
                        StockMovementType.sale
                        if delta < 0
                        else StockMovementType.restock
                    ),
                    "source": "hourly",
                    "snapshot_at": h.created_at,
                })

                prev_cache[key] = stock_after

            # -----------------------------------------------
            # bulk insert (NO COMMIT HERE)
            # -----------------------------------------------
            if rows:
                StockMovementRepo.bulk_insert_mappings_no_commit(db, rows)
                total_created += len(rows)

        # ======================================================
        # UPDATE CURSOR (ONCE)
        # ======================================================
        if last_processed_at:
            StockMovementRepo.update_cursor(
                db,
                cursor,
                last_hourly_processed_at=last_processed_at,
            )

        return total_created
    
    @staticmethod
    def build_hourly_movements_all(
        db: Session,
        *,
        chunk_size: int = 5000,
    ) -> int:
        """
        Обрабатывает ВСЕ provider_name.
        Возвращает общее количество созданных stock_movements.
        """

        total_created = 0

        provider_names = HourlyRepo.get_distinct_providers(db)
        if not provider_names:
            return 0

        for provider_name in provider_names:
            created = StockMovementService._build_for_provider(
                db,
                provider_name=provider_name,
                chunk_size=chunk_size,
            )
            total_created += created

        db.commit()
        return total_created
